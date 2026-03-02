"""
QR Code generation module.
Generates encrypted QR codes for students from an Excel master list.
"""
import os
import qrcode
from qrcode.image.styledpil import StyledPilImage
from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook
from crypto_utils import encrypt_qr_data, generate_data_hash
from config import QR_CODES_DIR


def generate_single_qr(student_data: dict, output_dir: str = None) -> str:
    """
    Generate a single QR code for a student.
    Returns the file path of the generated QR code.
    """
    if output_dir is None:
        output_dir = QR_CODES_DIR
    os.makedirs(output_dir, exist_ok=True)

    # Build payload
    payload = {
        'name': student_data['name'],
        'student_number': student_data['student_number'],
        'course': student_data['course'],
        'year': student_data['year'],
        'section': student_data['section'],
        'hash': generate_data_hash(student_data)
    }

    encrypted = encrypt_qr_data(payload)

    # Generate QR code
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=10,
        border=4,
    )
    qr.add_data(encrypted)
    qr.make(fit=True)

    qr_img = qr.make_image(fill_color="#1a1a2e", back_color="white").convert('RGB')

    # Create a labeled card
    card_width = max(qr_img.width, 400)
    card_height = qr_img.height + 120
    card = Image.new('RGB', (card_width, card_height), 'white')

    # Center QR on card
    qr_x = (card_width - qr_img.width) // 2
    card.paste(qr_img, (qr_x, 10))

    # Add text label below QR
    draw = ImageDraw.Draw(card)
    try:
        font = ImageFont.truetype("arial.ttf", 16)
        font_small = ImageFont.truetype("arial.ttf", 13)
    except (IOError, OSError):
        font = ImageFont.load_default()
        font_small = font

    label_y = qr_img.height + 15
    name_text = student_data['name']
    id_text = f"{student_data['student_number']} | {student_data['course']} {student_data['year']}-{student_data['section']}"

    # Center text
    name_bbox = draw.textbbox((0, 0), name_text, font=font)
    name_w = name_bbox[2] - name_bbox[0]
    draw.text(((card_width - name_w) // 2, label_y), name_text, fill="#1a1a2e", font=font)

    id_bbox = draw.textbbox((0, 0), id_text, font=font_small)
    id_w = id_bbox[2] - id_bbox[0]
    draw.text(((card_width - id_w) // 2, label_y + 25), id_text, fill="#555555", font=font_small)

    # Save
    safe_name = student_data['student_number'].replace(' ', '_').replace('/', '-')
    filename = f"QR_{safe_name}.png"
    filepath = os.path.join(output_dir, filename)
    card.save(filepath)

    return filepath


def batch_generate_from_excel(excel_path: str, output_dir: str = None) -> dict:
    """
    Generate QR codes for all students in an Excel file.
    Expected columns: Name, Student Number, Course, Year, Section
    Returns a dict with success count, error count, and error details.
    """
    if output_dir is None:
        output_dir = QR_CODES_DIR

    wb = load_workbook(excel_path)
    ws = wb.active

    # Find header row
    headers = {}
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value:
            headers[cell.value.strip().lower()] = col_idx

    # Map expected columns
    col_map = {}
    name_variants = ['name', 'full name', 'student name', 'fullname']
    number_variants = ['student number', 'student_number', 'studentnumber', 'id', 'student id', 'student_id']
    course_variants = ['course', 'program', 'degree']
    year_variants = ['year', 'year level', 'yr']
    section_variants = ['section', 'sec', 'sect']

    for key, variants in [('name', name_variants), ('student_number', number_variants),
                          ('course', course_variants), ('year', year_variants),
                          ('section', section_variants)]:
        for v in variants:
            if v in headers:
                col_map[key] = headers[v]
                break

    required = ['name', 'student_number', 'course', 'year', 'section']
    missing = [k for k in required if k not in col_map]
    if missing:
        return {
            'success': 0,
            'errors': 1,
            'error_details': [f"Missing required columns: {', '.join(missing)}. "
                              f"Found columns: {list(headers.keys())}"]
        }

    results = {'success': 0, 'errors': 0, 'error_details': [], 'files': []}

    for row in ws.iter_rows(min_row=2):
        try:
            student_data = {
                'name': str(row[col_map['name'] - 1].value or '').strip(),
                'student_number': str(row[col_map['student_number'] - 1].value or '').strip(),
                'course': str(row[col_map['course'] - 1].value or '').strip(),
                'year': str(row[col_map['year'] - 1].value or '').strip(),
                'section': str(row[col_map['section'] - 1].value or '').strip(),
            }

            if not student_data['name'] or not student_data['student_number']:
                continue  # Skip empty rows

            filepath = generate_single_qr(student_data, output_dir)
            results['files'].append(filepath)
            results['success'] += 1

        except Exception as e:
            results['errors'] += 1
            row_num = row[0].row if row else '?'
            results['error_details'].append(f"Row {row_num}: {str(e)}")

    wb.close()
    return results
