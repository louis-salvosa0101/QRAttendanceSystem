"""
Advanced QR-Based Attendance System - Main Flask Application
"""
import os
import io
import json
import zipfile
from datetime import datetime
from flask import (Flask, render_template, request, jsonify, redirect,
                   url_for, flash, send_from_directory, send_file)
from flask_login import login_required, login_user, logout_user, current_user
from werkzeug.utils import secure_filename
from config import (SECRET_KEY, QR_CODES_DIR, EXCEL_DIR, MASTER_LIST_DIR,
                    ATTENDANCE_LOG_FILE, FINE_LATE, FINE_ABSENT, FINE_PARTIAL,
                    LATE_THRESHOLD_MINUTES, ph_now, session_fine_value)
from crypto_utils import decrypt_qr_data
from qr_generator import generate_single_qr, batch_generate_from_excel
from session_manager import (create_session, get_session, get_active_sessions,
                              get_all_sessions, get_session_count, validate_session,
                              record_student_scan, close_session, clear_all_sessions,
                              delete_session, process_scan, get_session_row,
                              sync_scan_row_from_attendance)
from db import get_db, _cur
from excel_logger import (log_attendance, log_absent_students,
                           generate_summary_sheet,
                           get_attendance_records, get_session_stats,
                           create_sample_master_list, clear_attendance_records,
                           clear_session_records,
                           update_session_attendance_record, add_manual_attendance_record)
from student_registry import (register_student, register_students_bulk,
                               get_all_students, get_student,
                               get_students_by_filter, search_students_by_last_name,
                               delete_student,
                               update_student, clear_registry,
                               get_registry_stats)

from db import init_db
from auth import login_manager, authenticate, seed_default_admin, hash_password

app = Flask(__name__)


def _body_optional_int(data, key):
    """Parse JSON body field as int; treat None and blank strings as missing."""
    if not data:
        return None
    v = data.get(key)
    if v is None:
        return None
    if isinstance(v, str) and not str(v).strip():
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


@app.context_processor
def inject_config_fines():
    """Expose global fine defaults to templates (per-session amounts use these when unset)."""
    return {
        'FINE_LATE': FINE_LATE,
        'FINE_ABSENT': FINE_ABSENT,
        'FINE_PARTIAL': FINE_PARTIAL,
    }


with app.app_context():
    init_db()
    seed_default_admin()


@app.template_filter('fmt_dt')
def fmt_dt_filter(value):
    """Format an ISO-ish datetime string into 12-hour PH time like 'Apr 14, 2026 2:30 PM'."""
    if not value:
        return '—'
    try:
        from config import PH_TZ
        raw = str(value).replace('Z', '+00:00')
        dt = datetime.fromisoformat(raw)
        if dt.tzinfo is not None:
            dt = dt.astimezone(PH_TZ).replace(tzinfo=None)
        hour = dt.hour % 12 or 12
        return f"{dt.strftime('%b')} {dt.day}, {dt.year} {hour}:{dt.strftime('%M %p')}"
    except (ValueError, TypeError):
        fallback = str(value)[:16].replace('T', ' ')
        return fallback if fallback else '—'


app.secret_key = SECRET_KEY
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max upload

login_manager.init_app(app)

ALLOWED_EXTENSIONS = {'xlsx', 'xls'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# ─── AUTH ROUTES ──────────────────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Officer login page."""
    if current_user.is_authenticated:
        return redirect(url_for('index'))

    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')

        officer = authenticate(username, password)
        if officer:
            from flask import session as flask_session
            flask_session.pop('_officer_cache', None)
            login_user(officer)
            next_page = request.args.get('next')
            return redirect(next_page or url_for('index'))
        else:
            error = 'Invalid username or password.'

    return render_template('login.html', error=error)


@app.route('/logout')
@login_required
def logout():
    """Log out the current officer."""
    from flask import session as flask_session
    flask_session.pop('_officer_cache', None)
    logout_user()
    return redirect(url_for('login'))


@app.route('/api/change-password', methods=['POST'])
@login_required
def api_change_password():
    """Change the current officer's password."""
    import bcrypt
    data = request.get_json() or {}
    current_password = data.get('current_password', '')
    new_password = data.get('new_password', '')
    confirm_password = data.get('confirm_password', '')

    if not current_password or not new_password or not confirm_password:
        return jsonify({'success': False, 'message': 'All fields are required.'}), 400

    if new_password != confirm_password:
        return jsonify({'success': False, 'message': 'New passwords do not match.'}), 400

    if len(new_password) < 6:
        return jsonify({'success': False, 'message': 'New password must be at least 6 characters.'}), 400

    with get_db() as conn:
        cur = _cur(conn)
        cur.execute("SELECT password_hash FROM officers WHERE id = %s", (current_user.id,))
        row = cur.fetchone()

        if not row or not bcrypt.checkpw(current_password.encode('utf-8'),
                                          row['password_hash'].encode('utf-8')):
            return jsonify({'success': False, 'message': 'Current password is incorrect.'}), 403

        new_hash = hash_password(new_password)
        cur.execute("UPDATE officers SET password_hash = %s WHERE id = %s",
                    (new_hash, current_user.id))

    return jsonify({'success': True, 'message': 'Password changed successfully.'})


# ─── PAGES ───────────────────────────────────────────────────────────────

@app.route('/')
@login_required
def index():
    """Dashboard / Home page."""
    active_sessions = get_active_sessions()
    total_sessions = get_session_count()
    registry_stats = get_registry_stats()
    return render_template('index.html',
                           active_sessions=active_sessions,
                           total_sessions=total_sessions,
                           registry_stats=registry_stats)


@app.route('/scanner')
@login_required
def scanner_page():
    """QR Scanner page for teachers."""
    active_sessions = get_active_sessions()
    return render_template('scanner.html', active_sessions=active_sessions)


@app.route('/sessions')
@login_required
def sessions_page():
    """Session management page."""
    active_sessions = get_active_sessions()
    all_sessions = get_all_sessions()
    all_students = get_all_students()
    courses = sorted(set(s.get('course', '') for s in all_students if s.get('course')))
    return render_template('sessions.html',
                           active_sessions=active_sessions,
                           all_sessions=all_sessions,
                           courses=courses)


@app.route('/generate')
@login_required
def generate_page():
    """QR Code generation page."""
    qr_files = []
    if os.path.exists(QR_CODES_DIR):
        qr_files = [f for f in os.listdir(QR_CODES_DIR) if f.endswith('.png')]
    return render_template('generate.html', qr_files=qr_files)


@app.route('/records')
@login_required
def records_page():
    """Attendance records page."""
    session_id = request.args.get('session_id', '')
    student_number = request.args.get('student_number', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    records = get_attendance_records(
        session_id=session_id or None,
        student_number=student_number or None,
        date_from=date_from or None,
        date_to=date_to or None
    )

    all_sessions = get_all_sessions()

    return render_template('records.html',
                           records=records,
                           all_sessions=all_sessions,
                           filters={
                               'session_id': session_id,
                               'student_number': student_number,
                               'date_from': date_from,
                               'date_to': date_to
                           })


@app.route('/students')
@login_required
def students_page():
    """Student registry management page."""
    students = get_all_students()
    stats = get_registry_stats()
    return render_template('students.html', students=students, stats=stats)


@app.route('/officers')
@login_required
def officers_page():
    """Officer account management page."""
    with get_db() as conn:
        cur = _cur(conn)
        cur.execute("SELECT id, username, name, created_at, is_admin FROM officers ORDER BY id")
        officers = [dict(r) for r in cur.fetchall()]
    return render_template('officers.html', officers=officers)


@app.route('/students/<student_number>')
@login_required
def student_detail_page(student_number):
    """Individual student dashboard with attendance records and fines."""
    student = get_student(student_number)
    if not student or not student.get('student_number'):
        return render_template('404.html'), 404

    records = get_attendance_records(student_number=student_number)

    # Compute summary stats
    total_fines = 0
    absent_count = 0
    late_count = 0
    partial_count = 0
    time_in_count = 0
    time_out_count = 0
    sessions_set = set()
    fines_list = []

    for r in records:
        fine = r.get('fine') or 0
        total_fines += fine
        sess_id = r.get('session_id')
        if sess_id:
            sessions_set.add(sess_id)
        status = r.get('status', '')
        time_out_val = (r.get('time_out') or '').strip()
        if status == 'Absent':
            absent_count += 1
        elif status == 'Time In' and not time_out_val:
            time_in_count += 1
        elif status == 'Time Out' or time_out_val:
            time_out_count += 1
        if 'Partial' in status:
            partial_count += 1
        if fine > 0:
            late_count += 1
            fines_list.append({
                'id': r.get('id'),
                'date': r.get('datetime', ''),
                'session_id': sess_id,
                'status': status,
                'fine': fine,
                'reason': r.get('fine_reason', ''),
            })

    # Fetch manual fines and payments
    with get_db() as conn:
        cur = _cur(conn)
        cur.execute(
            "SELECT * FROM manual_fines WHERE student_number = %s ORDER BY created_at DESC",
            (student_number,)
        )
        manual_fines = cur.fetchall()

        cur.execute(
            "SELECT * FROM fine_payments WHERE student_number = %s ORDER BY created_at DESC",
            (student_number,)
        )
        payments = cur.fetchall()

    manual_fines_total = sum(f['amount'] for f in manual_fines)
    payments_total = sum(p['amount'] for p in payments)

    summary = {
        'total_sessions': len(sessions_set),
        'absent_count': absent_count,
        'late_count': late_count,
        'partial_count': partial_count,
        'time_in_count': time_in_count,
        'time_out_count': time_out_count,
        'attendance_fines': total_fines,
        'manual_fines': manual_fines_total,
        'total_fines': total_fines + manual_fines_total,
        'total_paid': payments_total,
        'balance': max(0, (total_fines + manual_fines_total) - payments_total),
    }

    return render_template('student_detail.html',
                           student=student,
                           records=records,
                           summary=summary,
                           fines_list=fines_list,
                           manual_fines=manual_fines,
                           payments=payments)


# ─── API ENDPOINTS ──────────────────────────────────────────────────────

@app.route('/api/session/create', methods=['POST'])
@login_required
def api_create_session():
    """Create a new attendance session."""
    data = request.get_json(silent=True) or {}
    subject = data.get('subject', '')
    teacher = data.get('teacher', '')
    notes = data.get('notes', '')
    duration_hours = data.get('duration_hours')
    required_course = data.get('required_course', '')
    required_year = data.get('required_year', [])
    required_section = data.get('required_section', '')
    scheduled_start = data.get('scheduled_start', '')
    fine_late = _body_optional_int(data, 'fine_late')
    fine_absent = _body_optional_int(data, 'fine_absent')
    fine_partial = _body_optional_int(data, 'fine_partial')
    late_threshold_minutes = _body_optional_int(data, 'late_threshold_minutes')

    session = create_session(
        subject=subject,
        teacher=teacher,
        notes=notes,
        duration_hours=duration_hours,
        required_course=required_course,
        required_year=required_year,
        required_section=required_section,
        scheduled_start=scheduled_start,
        fine_late=fine_late,
        fine_absent=fine_absent,
        fine_partial=fine_partial,
        late_threshold_minutes=late_threshold_minutes,
    )
    return jsonify({'success': True, 'session': session})


@app.route('/api/session/<session_id>/close', methods=['POST'])
@login_required
def api_close_session(session_id):
    """
    Close an attendance session.
    After closing, mark absent all registered students who were required but didn't scan.
    """
    session = get_session(session_id)
    if not session:
        return jsonify({'success': False, 'message': 'Session not found.'})

    success, message = close_session(session_id)
    if not success:
        return jsonify({'success': False, 'message': message})

    session = get_session(session_id)

    required_students = get_students_by_filter(
        course=session.get('required_course') or None,
        year=session.get('required_year') or None,
        section=session.get('required_section') or None,
    )

    absent_result = {'absent_logged': 0, 'partial_updated': 0}
    if required_students:
        absent_result = log_absent_students(session_id, session, required_students)

    from config import FINE_ABSENT as _FA, FINE_PARTIAL as _FP
    return jsonify({
        'success': True,
        'message': message,
        'absent_logged': absent_result.get('absent_logged', 0),
        'partial_updated': absent_result.get('partial_updated', 0),
        'fine_absent': session_fine_value(session, 'fine_absent', _FA),
        'fine_partial': session_fine_value(session, 'fine_partial', _FP),
    })


@app.route('/api/session/clear-history', methods=['DELETE'])
@login_required
def api_clear_session_history():
    """Clear all session history from database."""
    try:
        deleted = clear_all_sessions()
        return jsonify({
            'success': True,
            'message': f'Session history cleared ({deleted} sessions removed).' if deleted
            else 'No session history to clear.'
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/session/<session_id>', methods=['DELETE'])
@login_required
def api_delete_session(session_id):
    """Delete a specific session and its associated scans."""
    try:
        deleted = delete_session(session_id)
        return jsonify({
            'success': True,
            'message': f'Session {session_id} deleted.' if deleted
            else f'Session {session_id} not found.'
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/session/<session_id>/stats')
@login_required
def api_session_stats(session_id):
    """Get statistics for a session."""
    stats = get_session_stats(session_id)
    session = get_session(session_id)
    return jsonify({'success': True, 'stats': stats, 'session': session})


@app.route('/api/session/<session_id>/attendance/<int:record_id>', methods=['PATCH'])
@login_required
def api_patch_session_attendance(session_id, record_id):
    """Update one attendance row for a closed session (correct status, times, fines)."""
    data = request.get_json(silent=True) or {}
    allowed_keys = (
        'status', 'time_in', 'time_out', 'fine', 'fine_reason',
        'name', 'course', 'year', 'section',
    )
    fields = {k: data[k] for k in allowed_keys if k in data}
    if not fields:
        return jsonify({'success': False, 'message': 'No valid fields to update.'}), 400

    ok, msg, record = update_session_attendance_record(record_id, session_id, fields)
    if not ok:
        return jsonify({'success': False, 'message': msg}), 400

    sync_scan_row_from_attendance(
        session_id,
        str(record['student_number']),
        record['status'],
        record.get('time_in'),
        record.get('time_out'),
        int(record.get('fine') or 0),
        record.get('fine_reason') or '',
    )
    return jsonify({'success': True, 'message': msg, 'record': record})


@app.route('/api/session/<session_id>/attendance', methods=['POST'])
@login_required
def api_add_session_attendance(session_id):
    """Add an attendance row for a student with no row yet (closed session only)."""
    data = request.get_json(silent=True) or {}
    sn = (data.get('student_number') or '').strip()
    if not sn:
        return jsonify({'success': False, 'message': 'student_number is required.'}), 400

    student = get_student(sn)
    if not student or not student.get('student_number'):
        return jsonify({'success': False, 'message': 'Student not found in registry.'}), 404

    ok, msg, record = add_manual_attendance_record(
        session_id,
        student,
        data.get('status', 'Absent'),
        data.get('time_in'),
        data.get('time_out'),
        data.get('fine', 0),
        data.get('fine_reason', ''),
    )
    if not ok:
        status_code = 409 if 'already has' in (msg or '').lower() else 400
        return jsonify({'success': False, 'message': msg}), status_code

    sync_scan_row_from_attendance(
        session_id,
        str(record['student_number']),
        record['status'],
        record.get('time_in'),
        record.get('time_out'),
        int(record.get('fine') or 0),
        record.get('fine_reason') or '',
    )
    return jsonify({'success': True, 'message': msg, 'record': record})


@app.route('/api/session/<session_id>/attendance/bulk-absent-missing', methods=['POST'])
@login_required
def api_bulk_absent_missing(session_id):
    """
    For a closed session with required course/year/section, add Absent rows
    (with session absent fine) for each required student who has no attendance row yet.
    """
    session = get_session(session_id)
    if not session:
        return jsonify({'success': False, 'message': 'Session not found.'}), 404
    if session.get('is_active'):
        return jsonify({'success': False, 'message': 'Session must be closed.'}), 400

    req_c = session.get('required_course') or None
    req_y = session.get('required_year') or []
    if not isinstance(req_y, list):
        req_y = []
    req_s = session.get('required_section') or None
    if not req_c and not req_y and not req_s:
        return jsonify({
            'success': False,
            'message': 'This session has no required attendees filter. Use Add student to add individuals.',
        }), 400

    year_arg = req_y if len(req_y) > 0 else None
    required = get_students_by_filter(course=req_c, year=year_arg, section=req_s)
    if not required:
        return jsonify({'success': False, 'message': 'No students match this session’s required filter.'}), 400

    existing = {str(r['student_number']) for r in get_attendance_records(session_id=session_id)
                if r.get('student_number')}
    fine_absent = session_fine_value(session, 'fine_absent', FINE_ABSENT)
    default_reason = 'Absent — roster correction (manual bulk)'
    data = request.get_json(silent=True) or {}
    fine_reason = (data.get('fine_reason') or '').strip() or default_reason

    added = 0
    errors = []
    for stu in required:
        sn = str(stu.get('student_number', '')).strip()
        if not sn or sn in existing:
            continue
        ok, msg, record = add_manual_attendance_record(
            session_id, stu, 'Absent', None, None, fine_absent, fine_reason,
        )
        if ok and record:
            sync_scan_row_from_attendance(
                session_id, sn, record['status'],
                record.get('time_in'), record.get('time_out'),
                int(record.get('fine') or 0), record.get('fine_reason') or '',
            )
            existing.add(sn)
            added += 1
        else:
            errors.append(f'{sn}: {msg}')

    if added == 0 and errors:
        return jsonify({
            'success': False,
            'message': errors[0],
            'added': 0,
            'errors': errors[:10],
        }), 400

    msg = f'Added {added} absent record(s) for students missing from this session.'
    if errors:
        msg += f' ({len(errors)} skipped.)'
    return jsonify({
        'success': True,
        'message': msg,
        'added': added,
        'errors': errors[:20],
    })


@app.route('/api/students/search-by-last-name', methods=['GET'])
@login_required
def api_search_students_by_last_name():
    """Search students by last name (or name substring); optional session scope and exclusions."""
    q = request.args.get('q', '').strip()
    if not q:
        return jsonify({'success': True, 'students': []})

    try:
        lim = min(50, max(1, int(request.args.get('limit', 25))))
    except (TypeError, ValueError):
        lim = 25

    session_id = (request.args.get('session_id') or '').strip()
    exclude_in = request.args.get('exclude_in_session', '1').lower() in ('1', 'true', 'yes')

    allowed_numbers = None
    exclude_numbers = None

    if session_id:
        sess = get_session(session_id)
        if not sess:
            return jsonify({'success': False, 'message': 'Session not found.'}), 404
        rc = sess.get('required_course') or None
        ry = sess.get('required_year') or []
        if not isinstance(ry, list):
            ry = []
        rs = sess.get('required_section') or None
        if rc or ry or rs:
            pool = get_students_by_filter(
                course=rc,
                year=ry if len(ry) > 0 else None,
                section=rs,
            )
            allowed_numbers = [str(s['student_number']) for s in pool if s.get('student_number')]

        if exclude_in:
            exclude_numbers = list({
                str(r['student_number'])
                for r in get_attendance_records(session_id=session_id)
                if r.get('student_number')
            })

    students = search_students_by_last_name(
        q, lim,
        allowed_numbers=allowed_numbers,
        exclude_numbers=exclude_numbers,
    )
    return jsonify({'success': True, 'students': students})


@app.route('/api/scan', methods=['POST'])
@login_required
def api_scan_qr():
    """
    Process a scanned QR code.
    Uses a single pooled DB connection for the entire pipeline:
      validate session -> check filters -> register student -> record scan -> log attendance.
    """
    data = request.get_json()
    qr_content = data.get('qr_data', '')
    session_id = data.get('session_id', '')

    if not session_id:
        return jsonify({
            'success': False,
            'error': 'no_session',
            'message': 'No active session selected. Please select or create a session first.'
        })

    # Decrypt first (CPU-only, no DB needed)
    student_data = decrypt_qr_data(qr_content)
    if not student_data:
        return jsonify({
            'success': False,
            'error': 'invalid_qr',
            'message': 'Invalid or tampered QR code. Could not decrypt data.'
        })

    required_fields = ['name', 'student_number', 'course', 'year', 'section']
    for field in required_fields:
        if field not in student_data:
            return jsonify({
                'success': False,
                'error': 'invalid_qr',
                'message': f'QR code is missing required field: {field}'
            })

    # --- single DB connection for everything below ---
    with get_db() as conn:
        # 1. Lightweight session fetch (no scanned-students load)
        session = get_session_row(conn, session_id)
        if not session:
            return jsonify({
                'success': False,
                'error': 'invalid_session',
                'message': 'Session not found.'
            })

        if not session.get('is_active'):
            return jsonify({
                'success': False,
                'error': 'invalid_session',
                'message': 'Session has been closed.'
            })

        now = ph_now()
        expires_at = datetime.fromisoformat(session['expires_at'])
        if now >= expires_at:
            _cur(conn).execute(
                "UPDATE sessions SET is_active = 0 WHERE session_id = %s",
                (session_id,),
            )
            return jsonify({
                'success': False,
                'error': 'invalid_session',
                'message': 'Session has expired.'
            })

        # 2. Course / year / section filter check
        req_course = session.get('required_course') or ''
        req_year = session.get('required_year') or []
        req_section = session.get('required_section') or ''
        student_course = str(student_data.get('course', '')).strip()
        student_year = str(student_data.get('year', '')).strip()
        student_section = str(student_data.get('section', '')).strip()

        not_included_reasons = []
        if req_course and student_course != req_course:
            not_included_reasons.append(f"course ({student_course} != {req_course})")
        if req_year and isinstance(req_year, list) and len(req_year) > 0 and student_year not in req_year:
            year_labels = {'1': '1st', '2': '2nd', '3': '3rd', '4': '4th', '5': '5th'}
            allowed = ', '.join(year_labels.get(y, y) + ' year' for y in req_year)
            not_included_reasons.append(
                f"year level (you are {year_labels.get(student_year, student_year)} year; "
                f"session is for {allowed} only)"
            )
        if req_section and student_section != req_section:
            not_included_reasons.append(f"section ({student_section} != {req_section})")

        if not_included_reasons:
            return jsonify({
                'success': False,
                'error': 'not_included',
                'message': f"{student_data['name']} is not included in this session. "
                           f"This session is for {', '.join(not_included_reasons)}.",
                'student': student_data
            })

        # 3. Register / upsert student (reuses conn)
        register_student(student_data, conn=conn)

        # 4. Validate + record scan + get count (reuses conn)
        success, scan_msg, scan_type, fine, fine_reason, attendance_count, retry_after = \
            process_scan(conn, session_id, student_data['student_number'])

        if not success:
            err = 'cooldown' if retry_after is not None else 'duplicate'
            body = {
                'success': False,
                'error': err,
                'message': f"{student_data['name']} ({student_data['student_number']}) - {scan_msg}",
                'student': student_data,
            }
            if retry_after is not None:
                body['retry_after_seconds'] = retry_after
            return jsonify(body)

        status = 'Time In' if scan_type == 'time_in' else 'Time Out'

        # 5. Log attendance record (reuses conn)
        log_success = log_attendance(student_data, session_id, status=status,
                                     fine=fine, fine_reason=fine_reason, conn=conn)
        if not log_success:
            conn.rollback()
            return jsonify({
                'success': False,
                'error': 'log_error',
                'message': 'Failed to log attendance.'
            })

    fine_msg = f' | Fine: ₱{fine} ({fine_reason})' if fine else ''
    return jsonify({
        'success': True,
        'message': f"{student_data['name']} - {status} recorded!{fine_msg}",
        'student': student_data,
        'scan_type': scan_type,
        'status': status,
        'fine': fine,
        'fine_reason': fine_reason,
        'attendance_count': attendance_count
    })


# ─── STUDENT REGISTRY API ──────────────────────────────────────────────

@app.route('/api/students', methods=['GET'])
@login_required
def api_list_students():
    """List all registered students."""
    students = get_all_students()
    stats = get_registry_stats()
    return jsonify({'success': True, 'students': students, 'stats': stats})


@app.route('/api/students/register', methods=['POST'])
@login_required
def api_register_student():
    """Register a single student manually."""
    data = request.get_json()
    try:
        student_data = {
            'name': data.get('name', '').strip(),
            'student_number': data.get('student_number', '').strip(),
            'course': data.get('course', '').strip(),
            'year': data.get('year', '').strip(),
            'section': data.get('section', '').strip(),
        }
        if not student_data['name'] or not student_data['student_number']:
            return jsonify({'success': False, 'message': 'Name and Student Number are required.'})

        is_new = register_student(student_data)
        return jsonify({
            'success': True,
            'message': f"Student {'registered' if is_new else 'updated'} successfully.",
            'is_new': is_new
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/students/register-qr', methods=['POST'])
@login_required
def api_register_student_qr():
    """
    Register a student by scanning their QR code.
    Decrypts the QR payload and adds the student to the registry.
    No active session is needed -- this is registry-only.
    """
    data = request.get_json()
    qr_content = data.get('qr_data', '').strip()

    if not qr_content:
        return jsonify({'success': False, 'message': 'No QR data provided.'})

    student_data = decrypt_qr_data(qr_content)
    if not student_data:
        return jsonify({
            'success': False,
            'message': 'Invalid or tampered QR code. Could not read student data.'
        })

    required_fields = ['name', 'student_number', 'course', 'year', 'section']
    for field in required_fields:
        if field not in student_data:
            return jsonify({
                'success': False,
                'message': f'QR code is missing field: {field}'
            })

    is_new = register_student(student_data)
    return jsonify({
        'success': True,
        'message': f"{'Registered' if is_new else 'Updated'}: {student_data['name']} ({student_data['student_number']})",
        'student': student_data,
        'is_new': is_new
    })


@app.route('/api/students/import', methods=['POST'])
@login_required
def api_import_students():
    """Import students from an uploaded Excel file into the registry."""
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file uploaded.'})

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No file selected.'})

    if not allowed_file(file.filename):
        return jsonify({'success': False, 'message': 'Invalid file type. Please upload .xlsx or .xls file.'})

    try:
        from openpyxl import load_workbook
        filename = secure_filename(file.filename)
        upload_path = os.path.join(MASTER_LIST_DIR, filename)
        file.save(upload_path)

        wb = load_workbook(upload_path)
        ws = wb.active

        headers = {}
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value:
                headers[cell.value.strip().lower()] = col_idx

        col_map = {}
        name_variants = ['name', 'full name', 'student name', 'fullname']
        number_variants = ['student number', 'student_number', 'studentnumber', 'id', 'student id', 'student_id']
        course_variants = ['course', 'program', 'degree']
        year_variants = ['year', 'year level', 'yr']
        section_variants = ['section', 'sec', 'sect']

        for key, variants in [('name', name_variants), ('student_number', number_variants),
                               ('course', course_variants), ('year', year_variants),
                               ('section', section_variants)]:
            for v in variants:
                if v in headers:
                    col_map[key] = headers[v]
                    break

        required = ['name', 'student_number', 'course', 'year', 'section']
        missing = [k for k in required if k not in col_map]
        if missing:
            wb.close()
            return jsonify({
                'success': False,
                'message': f"Missing required columns: {', '.join(missing)}. Found: {list(headers.keys())}"
            })

        students = []
        for row in ws.iter_rows(min_row=2):
            name = str(row[col_map['name'] - 1].value or '').strip()
            student_number = str(row[col_map['student_number'] - 1].value or '').strip()
            if not name or not student_number:
                continue
            students.append({
                'name': name,
                'student_number': student_number,
                'course': str(row[col_map['course'] - 1].value or '').strip(),
                'year': str(row[col_map['year'] - 1].value or '').strip(),
                'section': str(row[col_map['section'] - 1].value or '').strip(),
            })
        wb.close()

        result = register_students_bulk(students)
        return jsonify({
            'success': True,
            'message': f"Import complete: {result['added']} new, {result['updated']} updated.",
            'added': result['added'],
            'updated': result['updated'],
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/students/<student_number>', methods=['PUT'])
@login_required
def api_update_student(student_number):
    """Update a student's profile information."""
    data = request.get_json() or {}
    success, message = update_student(
        student_number,
        name=data.get('name'),
        course=data.get('course'),
        year=data.get('year'),
        section=data.get('section'),
        new_student_number=data.get('new_student_number'),
    )
    new_sn = (data.get('new_student_number') or '').strip() or student_number
    return jsonify({'success': success, 'message': message,
                    'student_number': new_sn if success else student_number})


@app.route('/api/students/<student_number>', methods=['DELETE'])
@login_required
def api_delete_student(student_number):
    """Delete a student from the registry."""
    success = delete_student(student_number)
    return jsonify({
        'success': success,
        'message': 'Student deleted.' if success else 'Student not found.'
    })


@app.route('/api/students/clear', methods=['DELETE'])
@login_required
def api_clear_students():
    """Clear entire student registry."""
    count = clear_registry()
    return jsonify({'success': True, 'message': f'{count} students removed from registry.'})


# ─── OFFICER MANAGEMENT API ─────────────────────────────────────────────

@app.route('/api/officers', methods=['POST'])
@login_required
def api_create_officer():
    """Create a new officer account (admin only)."""
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'Only administrators can add officers.'}), 403

    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    username = (data.get('username') or '').strip()
    password = data.get('password', '')

    if not name or not username or not password:
        return jsonify({'success': False, 'message': 'Name, username, and password are required.'}), 400
    if len(password) < 6:
        return jsonify({'success': False, 'message': 'Password must be at least 6 characters.'}), 400

    with get_db() as conn:
        cur = _cur(conn)
        cur.execute("SELECT 1 FROM officers WHERE username = %s", (username,))
        if cur.fetchone():
            return jsonify({'success': False, 'message': f'Username "{username}" is already taken.'}), 409

        pw_hash = hash_password(password)
        now = ph_now().isoformat()
        cur.execute(
            "INSERT INTO officers (username, password_hash, name, created_at) VALUES (%s, %s, %s, %s)",
            (username, pw_hash, name, now),
        )

    return jsonify({'success': True, 'message': f'Officer "{name}" created successfully.'})


@app.route('/api/officers/<int:officer_id>', methods=['PUT'])
@login_required
def api_update_officer(officer_id):
    """Update an officer's name and/or username.
    Admins can edit anyone; regular officers can only edit themselves."""
    if not current_user.is_admin and officer_id != current_user.id:
        return jsonify({'success': False, 'message': 'You can only edit your own account.'}), 403

    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    username = (data.get('username') or '').strip()

    if not name or not username:
        return jsonify({'success': False, 'message': 'Name and username are required.'}), 400

    with get_db() as conn:
        cur = _cur(conn)
        cur.execute("SELECT id FROM officers WHERE id = %s", (officer_id,))
        if not cur.fetchone():
            return jsonify({'success': False, 'message': 'Officer not found.'}), 404

        cur.execute(
            "SELECT 1 FROM officers WHERE username = %s AND id != %s",
            (username, officer_id),
        )
        if cur.fetchone():
            return jsonify({'success': False, 'message': f'Username "{username}" is already taken.'}), 409

        cur.execute(
            "UPDATE officers SET name = %s, username = %s WHERE id = %s",
            (name, username, officer_id),
        )

    if officer_id == current_user.id:
        from flask import session as flask_session
        flask_session.pop('_officer_cache', None)

    return jsonify({'success': True, 'message': 'Officer updated successfully.'})


@app.route('/api/officers/<int:officer_id>/reset-password', methods=['POST'])
@login_required
def api_reset_officer_password(officer_id):
    """Reset another officer's password (admin only)."""
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'Only administrators can reset passwords.'}), 403

    data = request.get_json() or {}
    new_password = data.get('new_password', '')

    if not new_password or len(new_password) < 6:
        return jsonify({'success': False, 'message': 'Password must be at least 6 characters.'}), 400

    with get_db() as conn:
        cur = _cur(conn)
        cur.execute("SELECT id FROM officers WHERE id = %s", (officer_id,))
        if not cur.fetchone():
            return jsonify({'success': False, 'message': 'Officer not found.'}), 404

        pw_hash = hash_password(new_password)
        cur.execute("UPDATE officers SET password_hash = %s WHERE id = %s", (pw_hash, officer_id))

    return jsonify({'success': True, 'message': 'Password reset successfully.'})


@app.route('/api/officers/<int:officer_id>', methods=['DELETE'])
@login_required
def api_delete_officer(officer_id):
    """Delete an officer account (admin only, cannot delete yourself)."""
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'Only administrators can delete officers.'}), 403
    if officer_id == current_user.id:
        return jsonify({'success': False, 'message': 'You cannot delete your own account.'}), 403

    with get_db() as conn:
        cur = _cur(conn)
        cur.execute("DELETE FROM officers WHERE id = %s", (officer_id,))
        if cur.rowcount == 0:
            return jsonify({'success': False, 'message': 'Officer not found.'}), 404

    return jsonify({'success': True, 'message': 'Officer deleted successfully.'})


# ─── FINES & PAYMENTS API ───────────────────────────────────────────────

def _student_financial_snapshot(cur, student_number: str) -> dict:
    """Totals used for balance after fine/payment changes (same rules as profile page)."""
    cur.execute(
        "SELECT COALESCE(SUM(fine), 0) AS total FROM attendance_records WHERE student_number = %s",
        (student_number,),
    )
    attendance_fines = int(cur.fetchone()['total'] or 0)
    cur.execute(
        "SELECT COUNT(*) AS c FROM attendance_records WHERE student_number = %s AND COALESCE(fine, 0) > 0",
        (student_number,),
    )
    attendance_fine_rows = int(cur.fetchone()['c'] or 0)
    cur.execute(
        "SELECT COALESCE(SUM(amount), 0) AS total FROM manual_fines WHERE student_number = %s",
        (student_number,),
    )
    manual_fines = int(cur.fetchone()['total'] or 0)
    cur.execute(
        "SELECT COUNT(*) AS c FROM manual_fines WHERE student_number = %s",
        (student_number,),
    )
    manual_fine_rows = int(cur.fetchone()['c'] or 0)
    cur.execute(
        "SELECT COALESCE(SUM(amount), 0) AS total FROM fine_payments WHERE student_number = %s",
        (student_number,),
    )
    total_paid = int(cur.fetchone()['total'] or 0)
    total_fines = attendance_fines + manual_fines
    balance = max(0, total_fines - total_paid)
    return {
        'attendance_fines': attendance_fines,
        'attendance_fine_rows': attendance_fine_rows,
        'manual_fines': manual_fines,
        'manual_fine_rows': manual_fine_rows,
        'total_fines': total_fines,
        'total_paid': total_paid,
        'balance': balance,
    }


@app.route('/api/students/<student_number>/fines', methods=['POST'])
@login_required
def api_add_manual_fine(student_number):
    """Add a manual fine to a student."""
    student = get_student(student_number)
    if not student or not student.get('student_number'):
        return jsonify({'success': False, 'message': 'Student not found.'}), 404

    data = request.get_json() or {}
    amount = data.get('amount')
    reason = data.get('reason', '').strip()

    if not amount or not isinstance(amount, (int, float)) or amount <= 0:
        return jsonify({'success': False, 'message': 'A valid positive amount is required.'}), 400
    if not reason:
        return jsonify({'success': False, 'message': 'A reason is required.'}), 400

    now = ph_now().strftime('%Y-%m-%d %H:%M:%S')
    with get_db() as conn:
        cur = _cur(conn)
        cur.execute(
            "INSERT INTO manual_fines (student_number, amount, reason, created_at, created_by) "
            "VALUES (%s, %s, %s, %s, %s)",
            (student_number, int(amount), reason, now, current_user.name)
        )

    return jsonify({
        'success': True,
        'message': f'Fine of ₱{int(amount)} added to {student["name"]}.'
    })


@app.route('/api/students/<student_number>/fines/<int:fine_id>', methods=['DELETE'])
@login_required
def api_delete_manual_fine(student_number, fine_id):
    """Delete a manual fine."""
    with get_db() as conn:
        cur = _cur(conn)
        cur.execute(
            "DELETE FROM manual_fines WHERE id = %s AND student_number = %s",
            (fine_id, student_number)
        )
        deleted = cur.rowcount
        financial = _student_financial_snapshot(cur, student_number)

    return jsonify({
        'success': deleted > 0,
        'message': 'Fine removed.' if deleted else 'Fine not found.',
        'financial': financial,
    })


@app.route('/api/students/<student_number>/fines/clear', methods=['DELETE'])
@login_required
def api_clear_manual_fines(student_number):
    """Delete all manual fines for a student."""
    with get_db() as conn:
        cur = _cur(conn)
        cur.execute(
            "DELETE FROM manual_fines WHERE student_number = %s",
            (student_number,)
        )
        deleted = cur.rowcount
        financial = _student_financial_snapshot(cur, student_number)

    return jsonify({
        'success': True,
        'message': f'{deleted} manual fine(s) removed.' if deleted else 'No manual fines to remove.',
        'financial': financial,
    })


@app.route('/api/students/<student_number>/attendance-fines/<int:record_id>', methods=['DELETE'])
@login_required
def api_waive_attendance_fine(student_number, record_id):
    """Clear fine on one attendance record (student profile)."""
    student = get_student(student_number)
    if not student or not student.get('student_number'):
        return jsonify({'success': False, 'message': 'Student not found.'}), 404

    with get_db() as conn:
        cur = _cur(conn)
        cur.execute(
            """SELECT id, COALESCE(fine, 0) AS fine FROM attendance_records
               WHERE id = %s AND student_number = %s""",
            (record_id, student_number),
        )
        row = cur.fetchone()
        if not row:
            return jsonify({'success': False, 'message': 'Record not found.'}), 404
        prev_fine = int(row['fine'] or 0)
        cur.execute(
            """UPDATE attendance_records
               SET fine = 0,
                   fine_reason = CASE WHEN COALESCE(fine, 0) > 0 THEN 'Waived' ELSE fine_reason END
               WHERE id = %s AND student_number = %s""",
            (record_id, student_number),
        )
        financial = _student_financial_snapshot(cur, student_number)

    return jsonify({
        'success': True,
        'message': 'Attendance fine removed.' if prev_fine > 0 else 'No fine on this record.',
        'financial': financial,
    })


@app.route('/api/students/<student_number>/payments', methods=['POST'])
@login_required
def api_add_payment(student_number):
    """Record a fine payment for a student."""
    student = get_student(student_number)
    if not student or not student.get('student_number'):
        return jsonify({'success': False, 'message': 'Student not found.'}), 404

    data = request.get_json() or {}
    amount = data.get('amount')
    notes = data.get('notes', '').strip()

    if not amount or not isinstance(amount, (int, float)) or amount <= 0:
        return jsonify({'success': False, 'message': 'A valid positive amount is required.'}), 400

    now = ph_now().strftime('%Y-%m-%d %H:%M:%S')
    with get_db() as conn:
        cur = _cur(conn)
        cur.execute(
            "INSERT INTO fine_payments (student_number, amount, notes, created_at, created_by) "
            "VALUES (%s, %s, %s, %s, %s)",
            (student_number, int(amount), notes or None, now, current_user.name)
        )

    return jsonify({
        'success': True,
        'message': f'Payment of ₱{int(amount)} recorded for {student["name"]}.'
    })


@app.route('/api/students/<student_number>/reset', methods=['DELETE'])
@login_required
def api_reset_student_data(student_number):
    """Reset all data for a student: attendance records, manual fines, and payments."""
    student = get_student(student_number)
    if not student or not student.get('student_number'):
        return jsonify({'success': False, 'message': 'Student not found.'}), 404

    with get_db() as conn:
        cur = _cur(conn)
        cur.execute("DELETE FROM attendance_records WHERE student_number = %s", (student_number,))
        attendance_deleted = cur.rowcount
        cur.execute("DELETE FROM manual_fines WHERE student_number = %s", (student_number,))
        fines_deleted = cur.rowcount
        cur.execute("DELETE FROM fine_payments WHERE student_number = %s", (student_number,))
        payments_deleted = cur.rowcount
        cur.execute("DELETE FROM session_scans WHERE student_number = %s", (student_number,))

    return jsonify({
        'success': True,
        'message': (f'All data reset for {student["name"]}: '
                    f'{attendance_deleted} attendance record(s), '
                    f'{fines_deleted} fine(s), {payments_deleted} payment(s) removed.')
    })


@app.route('/api/students/<student_number>/payments/<int:payment_id>', methods=['DELETE'])
@login_required
def api_delete_payment(student_number, payment_id):
    """Delete a payment record."""
    with get_db() as conn:
        cur = _cur(conn)
        cur.execute(
            "DELETE FROM fine_payments WHERE id = %s AND student_number = %s",
            (payment_id, student_number)
        )
        deleted = cur.rowcount

    return jsonify({
        'success': deleted > 0,
        'message': 'Payment removed.' if deleted else 'Payment not found.'
    })


# ─── QR GENERATION API ──────────────────────────────────────────────────

@app.route('/api/generate/single', methods=['POST'])
@login_required
def api_generate_single():
    """Generate a single QR code and auto-register the student."""
    data = request.get_json()
    try:
        student_data = {
            'name': data.get('name', '').strip(),
            'student_number': data.get('student_number', '').strip(),
            'course': data.get('course', '').strip(),
            'year': data.get('year', '').strip(),
            'section': data.get('section', '').strip(),
        }

        if not student_data['name'] or not student_data['student_number']:
            return jsonify({'success': False, 'message': 'Name and Student Number are required.'})

        register_student(student_data)

        filepath = generate_single_qr(student_data)
        filename = os.path.basename(filepath)

        return jsonify({
            'success': True,
            'message': f'QR code generated for {student_data["name"]}',
            'filename': filename,
            'url': url_for('serve_qr', filename=filename)
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/generate/batch', methods=['POST'])
@login_required
def api_generate_batch():
    """Generate QR codes from an uploaded Excel file and register all students."""
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file uploaded.'})

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No file selected.'})

    if not allowed_file(file.filename):
        return jsonify({'success': False, 'message': 'Invalid file type. Please upload .xlsx or .xls file.'})

    try:
        filename = secure_filename(file.filename)
        upload_path = os.path.join(MASTER_LIST_DIR, filename)
        file.save(upload_path)

        results = batch_generate_from_excel(upload_path)

        from openpyxl import load_workbook as lw
        wb = lw(upload_path)
        ws = wb.active
        headers = {}
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value:
                headers[cell.value.strip().lower()] = col_idx

        col_map = {}
        for key, variants in [
            ('name', ['name', 'full name', 'student name', 'fullname']),
            ('student_number', ['student number', 'student_number', 'studentnumber', 'id', 'student id']),
            ('course', ['course', 'program', 'degree']),
            ('year', ['year', 'year level', 'yr']),
            ('section', ['section', 'sec', 'sect']),
        ]:
            for v in variants:
                if v in headers:
                    col_map[key] = headers[v]
                    break

        students = []
        if all(k in col_map for k in ['name', 'student_number', 'course', 'year', 'section']):
            for row in ws.iter_rows(min_row=2):
                name = str(row[col_map['name'] - 1].value or '').strip()
                student_number = str(row[col_map['student_number'] - 1].value or '').strip()
                if not name or not student_number:
                    continue
                students.append({
                    'name': name,
                    'student_number': student_number,
                    'course': str(row[col_map['course'] - 1].value or '').strip(),
                    'year': str(row[col_map['year'] - 1].value or '').strip(),
                    'section': str(row[col_map['section'] - 1].value or '').strip(),
                })
        wb.close()

        if students:
            register_students_bulk(students)

        return jsonify({
            'success': True,
            'message': f"Generated {results['success']} QR codes with {results['errors']} errors. {len(students)} students registered.",
            'results': {
                'success': results['success'],
                'errors': results['errors'],
                'error_details': results['error_details'],
                'registered': len(students),
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/generate/sample', methods=['POST'])
@login_required
def api_generate_sample():
    """Generate a sample master list Excel file."""
    try:
        filepath = create_sample_master_list()
        return jsonify({
            'success': True,
            'message': 'Sample master list created.',
            'filepath': filepath
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/records/summary', methods=['POST'])
@login_required
def api_generate_summary():
    """Generate the summary sheet in the attendance log."""
    try:
        success = generate_summary_sheet()
        return jsonify({
            'success': success,
            'message': 'Summary sheet generated successfully.' if success else 'Failed to generate summary.'
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/records/student-fines-summary', methods=['GET'])
@login_required
def api_student_fines_summary():
    """Download Excel summary of fines for every registered student (includes zero sessions).

    Optional query: year — e.g. ?year=1 limits the export to students with that year level.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    all_students = get_all_students()
    year_filter = (request.args.get('year') or '').strip()
    if year_filter:
        all_students = [
            st for st in all_students
            if str(st.get('year') or '').strip() == year_filter
        ]
    records = get_attendance_records()

    student_data = {}
    for st in all_students:
        sn = st['student_number']
        student_data[sn] = {
            'name': st.get('name') or '',
            'student_number': sn,
            'course': st.get('course') or '',
            'year': st.get('year') or '',
            'section': st.get('section') or '',
            'total_fines': 0, 'absent': 0, 'late': 0, 'present': 0,
        }

    for r in records:
        sn = r.get('student_number')
        if not sn or sn not in student_data:
            continue
        sd = student_data[sn]
        fine_val = r['fine'] if r['fine'] else 0
        sd['total_fines'] += fine_val
        status = r.get('status', '')
        if status == 'Absent':
            sd['absent'] += 1
        elif status in ('Late', 'Partial (No Time Out)') or (
            status in ('Time In', 'Time Out') and fine_val > 0
        ):
            sd['late'] += 1
        else:
            sd['present'] += 1

    manual_fines_map = {}
    payments_map = {}
    with get_db() as conn:
        cur = _cur(conn)
        cur.execute("SELECT student_number, SUM(amount) AS total FROM manual_fines GROUP BY student_number")
        for row in cur.fetchall():
            manual_fines_map[row['student_number']] = row['total'] or 0
        cur.execute("SELECT student_number, SUM(amount) AS total FROM fine_payments GROUP BY student_number")
        for row in cur.fetchall():
            payments_map[row['student_number']] = row['total'] or 0

    for sn, sd in student_data.items():
        sd['manual_fines'] = manual_fines_map.get(sn, 0)
        sd['total_paid'] = payments_map.get(sn, 0)
        sd['grand_fines'] = sd['total_fines'] + sd['manual_fines']
        sd['balance'] = max(0, sd['grand_fines'] - sd['total_paid'])

    wb = Workbook()
    summary_sheet = wb.active
    summary_sheet.title = 'Summary'

    sum_headers = ['Name', 'Student Number', 'Course', 'Year', 'Section',
                   'Present', 'Late', 'Absent', 'Attendance Fines (PHP)',
                   'Manual Fines (PHP)', 'Total Fines (PHP)',
                   'Total Paid (PHP)', 'Balance (PHP)']
    summary_sheet.append(sum_headers)
    sum_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    sum_font = Font(color="FFFFFF", bold=True)
    for cell in summary_sheet[1]:
        cell.fill = sum_fill
        cell.font = sum_font
        cell.alignment = Alignment(horizontal='center')

    fine_fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
    fine_font = Font(bold=True, color='721c24')
    paid_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
    paid_font = Font(bold=True, color='155724')
    balance_fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
    balance_font = Font(bold=True, color='856404')

    if not student_data:
        summary_sheet.append(['No students registered.', '', '', '', '', '', '', '', '', '', '', '', ''])

    for sd in sorted(student_data.values(), key=lambda x: (x.get('name') or '', x.get('student_number') or '')):
        summary_sheet.append([
            sd['name'], sd['student_number'], sd['course'], sd['year'], sd['section'],
            sd['present'], sd['late'], sd['absent'],
            sd['total_fines'], sd['manual_fines'], sd['grand_fines'],
            sd['total_paid'], sd['balance'],
        ])
        row_num = summary_sheet.max_row
        fines_cell = summary_sheet.cell(row=row_num, column=11)
        if sd['grand_fines'] > 0:
            fines_cell.fill = fine_fill
            fines_cell.font = fine_font
        paid_cell = summary_sheet.cell(row=row_num, column=12)
        if sd['total_paid'] > 0:
            paid_cell.fill = paid_fill
            paid_cell.font = paid_font
        balance_cell = summary_sheet.cell(row=row_num, column=13)
        if sd['balance'] > 0:
            balance_cell.fill = balance_fill
            balance_cell.font = balance_font
        elif sd['grand_fines'] > 0 and sd['balance'] == 0:
            balance_cell.fill = paid_fill
            balance_cell.font = paid_font
            balance_cell.value = 'PAID'

    grand_total_fines = sum(sd['grand_fines'] for sd in student_data.values())
    grand_total_paid = sum(sd['total_paid'] for sd in student_data.values())
    grand_balance = max(0, grand_total_fines - grand_total_paid)

    grand_row = summary_sheet.max_row + 2
    summary_sheet.cell(row=grand_row, column=10, value='GRAND TOTAL:').font = Font(bold=True, size=12)
    summary_sheet.cell(row=grand_row, column=10).alignment = Alignment(horizontal='right')
    summary_sheet.cell(row=grand_row, column=11, value=grand_total_fines).font = Font(bold=True, size=13, color='C00000')
    summary_sheet.cell(row=grand_row, column=11).alignment = Alignment(horizontal='center')
    summary_sheet.cell(row=grand_row, column=12, value=grand_total_paid).font = Font(bold=True, size=13, color='155724')
    summary_sheet.cell(row=grand_row, column=12).alignment = Alignment(horizontal='center')
    summary_sheet.cell(row=grand_row, column=13, value=grand_balance).font = Font(bold=True, size=13, color='856404')
    summary_sheet.cell(row=grand_row, column=13).alignment = Alignment(horizontal='center')

    sum_widths = [30, 20, 15, 8, 10, 10, 10, 10, 20, 18, 18, 18, 18]
    for i, w in enumerate(sum_widths, 1):
        col_letter = get_column_letter(i)
        summary_sheet.column_dimensions[col_letter].width = w

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    fname = 'student_fines_summary.xlsx'
    if year_filter:
        safe = ''.join(c for c in year_filter if c.isalnum() or c in ('-', '_'))[:20] or 'year'
        fname = f'student_fines_summary_year_{safe}.xlsx'
    return send_file(
        out,
        as_attachment=True,
        download_name=fname,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/api/records/download')
@login_required
def api_download_records():
    """Download the attendance records as Excel (supports filters)."""
    session_id = request.args.get('session_id', '')
    student_number = request.args.get('student_number', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    records = get_attendance_records(
        session_id=session_id or None,
        student_number=student_number or None,
        date_from=date_from or None,
        date_to=date_to or None
    )

    # Build session_id -> session name lookup
    all_sessions_list = get_all_sessions()
    session_name_map = {}
    for s in all_sessions_list:
        sid = s.get('session_id', '')
        label = s.get('subject') or sid
        session_name_map[sid] = label

    if not records:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Records"
        ws.append([
            'Date & Time', 'Name', 'Student Number', 'Course', 'Year', 'Section', 'Session',
            'Time In', 'Time Out', 'Status', 'Fine', 'Fine Reason',
        ])
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return send_file(out, as_attachment=True, download_name='attendance_export.xlsx',
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    ws = wb.active
    if session_id:
        stitle = session_name_map.get(session_id, session_id) or 'Records'
        ws.title = (stitle[:31] if len(stitle) > 31 else stitle) or 'Records'
    else:
        ws.title = 'Records'

    headers = [
        'Date & Time', 'Name', 'Student Number', 'Course', 'Year', 'Section', 'Session',
        'Time In', 'Time Out', 'Status', 'Fine', 'Fine Reason',
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    total_fines = 0
    for r in records:
        fine_val = r['fine'] if r['fine'] else 0
        total_fines += fine_val
        ws.append([
            r['datetime'],
            r['name'],
            r['student_number'],
            r['course'],
            r['year'],
            r['section'],
            session_name_map.get(r['session_id'], r['session_id']),
            r.get('time_in') or '',
            r.get('time_out') or '',
            r['status'],
            fine_val,
            r.get('fine_reason', '')
        ])

    ws.append([])
    r = ws.max_row + 1
    lbl = Font(bold=True, size=12)
    val = Font(bold=True, size=12, color='C00000')
    if session_id:
        sess_label = session_name_map.get(session_id, session_id)
        c = ws.cell(row=r, column=1, value=f'Session: {sess_label}')
        c.font = Font(bold=True, size=12)
        r += 1
    footer_label_col = len(headers) - 1
    footer_val_col = len(headers)
    ws.cell(row=r, column=footer_label_col, value='Total rows (this export):').font = lbl
    ws.cell(row=r, column=footer_label_col).alignment = Alignment(horizontal='right')
    ws.cell(row=r, column=footer_val_col, value=len(records)).font = lbl
    ws.cell(row=r, column=footer_val_col).alignment = Alignment(horizontal='center')
    r += 1
    ws.cell(row=r, column=footer_label_col, value='Total fines (PHP):').font = lbl
    ws.cell(row=r, column=footer_label_col).alignment = Alignment(horizontal='right')
    ws.cell(row=r, column=footer_val_col, value=total_fines).font = val
    ws.cell(row=r, column=footer_val_col).alignment = Alignment(horizontal='center')

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    filename = "attendance_export.xlsx"
    if session_id:
        safe_name = session_name_map.get(session_id, session_id)
        safe_name = "".join(c if c.isalnum() or c in (' ', '-', '_') else '_' for c in safe_name).strip()
        filename = f"{safe_name}_records.xlsx"

    return send_file(
        out,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/api/records/reset', methods=['DELETE'])
@login_required
def api_reset_records():
    """Reset all attendance records by clearing the database."""
    try:
        deleted = clear_attendance_records()
        return jsonify({
            'success': True,
            'message': f'Attendance records reset ({deleted} records removed).' if deleted
            else 'No records to reset.'
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/records/session/<session_id>', methods=['DELETE'])
@login_required
def api_delete_session_records(session_id):
    """Delete all attendance records for a specific session."""
    try:
        deleted = clear_session_records(session_id)
        return jsonify({
            'success': True,
            'message': f'{deleted} record(s) removed from session {session_id}.' if deleted
            else 'No records found for this session.'
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/qrcodes/list')
@login_required
def api_list_qrcodes():
    """List all generated QR code files."""
    qr_files = []
    if os.path.exists(QR_CODES_DIR):
        qr_files = [f for f in os.listdir(QR_CODES_DIR) if f.endswith('.png')]
    return jsonify({'success': True, 'files': qr_files})


@app.route('/api/qrcodes/clear', methods=['DELETE'])
@login_required
def api_clear_qrcodes():
    """Delete all generated QR code files."""
    try:
        deleted = 0
        if os.path.exists(QR_CODES_DIR):
            for f in os.listdir(QR_CODES_DIR):
                if f.endswith('.png'):
                    os.remove(os.path.join(QR_CODES_DIR, f))
                    deleted += 1
        return jsonify({
            'success': True,
            'message': f'{deleted} QR code(s) deleted.',
            'deleted': deleted
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/qrcodes/download-all')
@login_required
def api_download_all_qrcodes():
    """Download all generated QR codes as a single ZIP file."""
    try:
        qr_files = []
        if os.path.exists(QR_CODES_DIR):
            qr_files = [f for f in os.listdir(QR_CODES_DIR) if f.endswith('.png')]

        if not qr_files:
            return jsonify({'success': False, 'message': 'No QR codes to download.'}), 404

        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for filename in qr_files:
                filepath = os.path.join(QR_CODES_DIR, filename)
                zf.write(filepath, filename)

        zip_buffer.seek(0)
        return send_file(
            zip_buffer,
            mimetype='application/zip',
            as_attachment=True,
            download_name='qr_codes_all.zip'
        )
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ─── STATIC FILE SERVING ────────────────────────────────────────────────

@app.route('/qrcodes/<filename>')
@login_required
def serve_qr(filename):
    """Serve generated QR code images."""
    return send_from_directory(QR_CODES_DIR, filename)


# ─── ERROR HANDLERS ─────────────────────────────────────────────────────

@app.errorhandler(404)
def not_found(e):
    return render_template('404.html'), 404


@app.errorhandler(500)
def server_error(e):
    return jsonify({'error': 'Internal server error'}), 500


# ─── RUN ─────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    print("\n" + "=" * 60)
    print("  QR Attendance System")
    print(f"  Local URL:   http://127.0.0.1:{port}")
    print("=" * 60)

    use_ngrok = os.environ.get('USE_NGROK', '0') == '1'
    if use_ngrok:
        try:
            from pyngrok import ngrok
            public_url = ngrok.connect(port, "http").public_url
            print(f"  Ngrok URL:   {public_url}")
        except Exception as e:
            print(f"  [!] ngrok failed: {e}")

    app.run(debug=True, host='0.0.0.0', port=port, use_reloader=False)
