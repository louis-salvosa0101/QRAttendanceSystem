"""
Attendance logging module.
Stores records in PostgreSQL (Supabase); provides Excel export for reports.
"""
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import (ATTENDANCE_LOG_FILE, EXCEL_DIR, FINE_ABSENT, FINE_PARTIAL,
                     FINE_LATE, PH_TZ, ph_now, session_fine_value)
from db import get_db, _cur


# Style constants for Excel export
HEADER_FONT = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid')
SUCCESS_FILL = PatternFill(start_color='d4edda', end_color='d4edda', fill_type='solid')
TIMEOUT_FILL = PatternFill(start_color='cce5ff', end_color='cce5ff', fill_type='solid')
ABSENT_FILL = PatternFill(start_color='f8d7da', end_color='f8d7da', fill_type='solid')
LATE_FILL = PatternFill(start_color='fff3cd', end_color='fff3cd', fill_type='solid')
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin', color='cccccc'),
    right=Side(style='thin', color='cccccc'),
    top=Side(style='thin', color='cccccc'),
    bottom=Side(style='thin', color='cccccc'),
)

ATTENDANCE_HEADERS = [
    'Date & Time', 'Student Name', 'Student Number', 'Course', 'Year', 'Section',
    'Session ID', 'Time In', 'Time Out', 'Status', 'Fine (PHP)', 'Fine Reason',
]
COLUMN_WIDTHS = [22, 30, 20, 15, 8, 10, 15, 18, 18, 15, 12, 40]


def _style_header_row(ws, headers, widths):
    """Apply styling to the header row."""
    for col_idx, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _get_session_sheet_name(session_id: str) -> str:
    """Generate a sheet name for a session."""
    today = ph_now().strftime('%Y-%m-%d')
    name = f"{today}_{session_id}"
    return name[:31]


def log_attendance(student_data: dict, session_id: str, status: str = "Present",
                   fine: int = 0, fine_reason: str = '', conn=None) -> bool:
    """
    Log attendance to PostgreSQL: Time In inserts one row; Time Out updates that row.
    If *conn* is provided the caller's connection is reused (no new pool checkout).
    """
    now = ph_now().strftime('%Y-%m-%d %H:%M:%S')
    sn = student_data.get('student_number', '')

    def _run(cur):
        if status == 'Time Out':
            cur.execute(
                """UPDATE attendance_records
                   SET time_out = %s, status = 'Time Out'
                   WHERE session_id = %s AND student_number = %s
                     AND status = 'Time In'
                     AND (time_out IS NULL OR TRIM(COALESCE(time_out, '')) = '')""",
                (now, session_id, sn),
            )
            return cur.rowcount > 0
        time_in_val = now if status == 'Time In' else None
        cur.execute(
            """INSERT INTO attendance_records
               (recorded_at, name, student_number, course, year, section, session_id,
                status, fine, fine_reason, time_in, time_out)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
            (
                now,
                student_data.get('name', ''),
                sn,
                student_data.get('course', ''),
                student_data.get('year', ''),
                student_data.get('section', ''),
                session_id,
                status,
                fine or 0,
                fine_reason or '',
                time_in_val,
                None,
            ),
        )
        return True

    try:
        if conn is not None:
            return _run(_cur(conn))
        with get_db() as c:
            return _run(_cur(c))
    except Exception as e:
        print(f"Error logging attendance: {e}")
        return False


def log_absent_students(session_id: str, session_data: dict,
                        required_students: list) -> dict:
    """
    After a session closes, log all registered students who did NOT scan as Absent.
    Update any student who only Time In (no Time Out) to Partial with FINE_PARTIAL.
    Uses per-session fine amounts when available, falling back to global config.
    Returns dict with counts: absent_logged, partial_updated
    """
    scanned = session_data.get('scanned_students', {})
    s_fine_absent = session_fine_value(session_data, 'fine_absent', FINE_ABSENT)
    s_fine_partial = session_fine_value(session_data, 'fine_partial', FINE_PARTIAL)
    result = {'absent_logged': 0, 'partial_updated': 0}
    now = ph_now().strftime('%Y-%m-%d %H:%M:%S')

    with get_db() as conn:
        cur = _cur(conn)

        all_student_numbers = [str(s.get('student_number', '')) for s in required_students]
        cur.execute(
            "SELECT DISTINCT student_number FROM attendance_records WHERE session_id = %s AND student_number = ANY(%s)",
            (session_id, all_student_numbers),
        )
        already_logged = {r['student_number'] for r in cur.fetchall()}

        for student in required_students:
            student_number = str(student.get('student_number', ''))
            scan_info = scanned.get(student_number, {})

            if student_number in already_logged and not scan_info:
                continue

            if not scan_info:
                cur.execute(
                    """INSERT INTO attendance_records
                       (recorded_at, name, student_number, course, year, section, session_id,
                        status, fine, fine_reason, time_in, time_out)
                       VALUES (%s, %s, %s, %s, %s, %s, %s, 'Absent', %s, %s, NULL, NULL)""",
                    (now, student.get('name', ''), student_number, student.get('course', ''),
                     student.get('year', ''), student.get('section', ''),
                     session_id, s_fine_absent, 'Absent - did not scan QR')
                )
                result['absent_logged'] += 1

            elif scan_info.get('status') == 'in':
                fine_reason = 'No Time Out recorded (partial scan) - considered late'

                cur.execute(
                    """UPDATE attendance_records
                       SET status = 'Partial (No Time Out)', fine = %s, fine_reason = %s
                       WHERE session_id = %s AND student_number = %s AND status = 'Time In'
                         AND (time_out IS NULL OR TRIM(COALESCE(time_out, '')) = '')""",
                    (s_fine_partial, fine_reason, session_id, student_number)
                )
                if cur.rowcount > 0:
                    result['partial_updated'] += 1
                else:
                    cur.execute(
                        """INSERT INTO attendance_records
                           (recorded_at, name, student_number, course, year, section, session_id,
                            status, fine, fine_reason, time_in, time_out)
                           VALUES (%s, %s, %s, %s, %s, %s, %s, 'Partial (No Time Out)', %s, %s, NULL, NULL)""",
                        (now, student.get('name', ''), student_number, student.get('course', ''),
                         student.get('year', ''), student.get('section', ''),
                         session_id, s_fine_partial, fine_reason),
                    )
                    result['partial_updated'] += 1

    return result


def generate_summary_sheet(filepath: str = None) -> bool:
    """
    Generate a summary Excel sheet with total attendance per student.
    Reads from PostgreSQL and writes to Excel file.
    """
    if filepath is None:
        filepath = ATTENDANCE_LOG_FILE

    records = get_attendance_records()
    student_records = {}

    for r in records:
        student_num = str(r.get('student_number', ''))
        if not student_num:
            continue
        if student_num not in student_records:
            student_records[student_num] = {
                'name': r.get('name'),
                'student_number': student_num,
                'course': r.get('course'),
                'year': r.get('year'),
                'section': r.get('section'),
                'total_sessions': 0,
                'sessions_attended': [],
                'total_fines': 0,
                'absent_count': 0,
                'late_count': 0,
            }
        sess_id = r.get('session_id')
        if sess_id and sess_id not in student_records[student_num]['sessions_attended']:
            student_records[student_num]['sessions_attended'].append(sess_id)
            student_records[student_num]['total_sessions'] += 1
        fine = r.get('fine') or 0
        student_records[student_num]['total_fines'] += fine
        if r.get('status') == 'Absent':
            student_records[student_num]['absent_count'] += 1
        elif r.get('status') in ('Late', 'Time In', 'Time Out', 'Partial (No Time Out)') and fine:
            student_records[student_num]['late_count'] += 1

    # Fetch manual fines and payments per student
    manual_fines_map = {}
    payments_map = {}
    with get_db() as conn:
        cur = _cur(conn)
        cur.execute("SELECT student_number, SUM(amount) AS total FROM manual_fines GROUP BY student_number")
        for row in cur.fetchall():
            manual_fines_map[row['student_number']] = row['total'] or 0
        cur.execute("SELECT student_number, SUM(amount) AS total FROM fine_payments GROUP BY student_number")
        for row in cur.fetchall():
            payments_map[row['student_number']] = row['total'] or 0

    for sn, data in student_records.items():
        data['manual_fines'] = manual_fines_map.get(sn, 0)
        data['total_paid'] = payments_map.get(sn, 0)
        data['grand_fines'] = data['total_fines'] + data['manual_fines']
        data['balance'] = max(0, data['grand_fines'] - data['total_paid'])

    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    summary_headers = [
        'Student Name', 'Student Number', 'Course', 'Year', 'Section',
        'Total Sessions', 'Sessions Attended', 'Absent Count',
        'Late Count', 'Attendance Fines (PHP)', 'Manual Fines (PHP)',
        'Total Fines (PHP)', 'Total Paid (PHP)', 'Balance (PHP)',
        'Sessions List'
    ]
    summary_widths = [30, 20, 15, 8, 10, 16, 18, 14, 12, 20, 18, 18, 18, 18, 50]

    PAID_FILL = PatternFill(start_color='d4edda', end_color='d4edda', fill_type='solid')

    try:
        ws = wb.create_sheet(title='Summary', index=0)
        _style_header_row(ws, summary_headers, summary_widths)

        for idx, (student_num, data) in enumerate(sorted(student_records.items()), 2):
            row_data = [
                data['name'], data['student_number'], data['course'], data['year'], data['section'],
                data['total_sessions'], data['total_sessions'], data['absent_count'], data['late_count'],
                data['total_fines'], data['manual_fines'], data['grand_fines'],
                data['total_paid'], data['balance'],
                ', '.join(str(s) for s in data['sessions_attended'])
            ]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=idx, column=col_idx, value=value)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if col_idx == 12 and value and value > 0:
                    cell.fill = ABSENT_FILL
                    cell.font = Font(name='Calibri', bold=True, color='721c24')
                elif col_idx == 13 and value and value > 0:
                    cell.fill = PAID_FILL
                    cell.font = Font(name='Calibri', bold=True, color='155724')
                elif col_idx == 14:
                    if value and value > 0:
                        cell.fill = LATE_FILL
                        cell.font = Font(name='Calibri', bold=True, color='856404')
                    elif data['grand_fines'] > 0 and data['balance'] == 0:
                        cell.value = 'PAID'
                        cell.fill = PAID_FILL
                        cell.font = Font(name='Calibri', bold=True, color='155724')

        wb.save(filepath)
        return True
    except Exception as e:
        print(f"Error generating summary: {e}")
        return False
    finally:
        wb.close()


def get_attendance_records(session_id: str = None, student_number: str = None,
                           date_from: str = None, date_to: str = None,
                           status: str = None, course: str = None,
                           year: str = None, section: str = None) -> list:
    """
    Retrieve attendance records from PostgreSQL with optional filters.
    """
    params = []
    conditions = []
    if session_id:
        conditions.append("ar.session_id = %s")
        params.append(session_id)
    if student_number:
        conditions.append("ar.student_number = %s")
        params.append(student_number)
    if date_from:
        conditions.append("ar.recorded_at >= %s")
        params.append(date_from + " 00:00:00")
    if date_to:
        conditions.append("ar.recorded_at <= %s")
        params.append(date_to + " 23:59:59")
    if status:
        st = status.strip()
        # "Late" is not usually stored as status: lateness is a late fine on Time In/Out rows.
        if st == "Late":
            # Session lateness is stored as Time In/Out with fine_reason "Late by … min…", not status='Late'.
            # Prefix matches session_manager (excludes partial "considered late" wording).
            conditions.append(
                """(
    TRIM(COALESCE(ar.status, '')) = 'Late'
    OR (
        COALESCE(ar.fine, 0) > 0
        AND ar.fine_reason IS NOT NULL
        AND ar.fine_reason ILIKE %s
    )
)"""
            )
            params.append("late by%")
        else:
            conditions.append("TRIM(COALESCE(ar.status, '')) = %s")
            params.append(st)
    _ilike_contains('ar.course', course, params, conditions)
    if year:
        conditions.append("ar.year = %s")
        params.append(year.strip())
    _ilike_contains('ar.section', section, params, conditions)

    where = " AND ".join(conditions) if conditions else "1=1"
    query = (
        "SELECT ar.id, ar.recorded_at AS datetime, ar.name, ar.student_number, ar.course, "
        "ar.year, ar.section, ar.session_id, ar.status, ar.fine, ar.fine_reason, ar.time_in, ar.time_out, "
        "s.subject AS session_subject, s.notes AS session_notes "
        "FROM attendance_records ar "
        "LEFT JOIN sessions s ON s.session_id = ar.session_id "
        f"WHERE {where} ORDER BY ar.recorded_at"
    )

    with get_db() as conn:
        cur = _cur(conn)
        cur.execute(query, params)
        rows = cur.fetchall()

    return [dict(r) for r in rows]


ALLOWED_EDIT_STATUSES = frozenset({'Time In', 'Time Out', 'Absent', 'Partial (No Time Out)'})


def _ilike_contains(column_sql: str, raw: str, params: list, conditions: list) -> None:
    """Add ILIKE %…% with ESCAPE so user % and _ are literal."""
    if not raw or not str(raw).strip():
        return
    frag = (
        str(raw).strip()
        .replace('\\', '\\\\')
        .replace('%', '\\%')
        .replace('_', '\\_')
    )
    conditions.append(f"{column_sql} ILIKE %s ESCAPE '\\'")
    params.append(f'%{frag}%')


def _dt_empty(value) -> bool:
    return value is None or (isinstance(value, str) and not str(value).strip())


def normalize_attendance_datetime(value) -> str | None:
    """Normalize UI/API datetime strings to 'YYYY-MM-DD HH:MM:SS' for storage."""
    if _dt_empty(value):
        return None
    raw = str(value).strip().replace('Z', '+00:00')
    if 'T' in raw and len(raw) == 16 and raw.count(':') == 1:
        raw = raw + ':00'
    try:
        dt = datetime.fromisoformat(raw.replace(' ', 'T'))
        if dt.tzinfo is not None:
            dt = dt.astimezone(PH_TZ).replace(tzinfo=None)
        return dt.strftime('%Y-%m-%d %H:%M:%S')
    except (ValueError, TypeError):
        return str(value).strip()


def assert_session_closed_for_attendance_edit(session_id: str) -> tuple[bool, str]:
    """Return (True, '') if session exists and is inactive; else (False, message)."""
    with get_db() as conn:
        cur = _cur(conn)
        cur.execute("SELECT is_active FROM sessions WHERE session_id = %s", (session_id,))
        row = cur.fetchone()
    if not row:
        return False, "Session not found."
    if int(row.get('is_active') or 0) != 0:
        return False, "Attendance can only be edited after the session is closed."
    return True, ""


def _validate_attendance_edit_payload(status: str, time_in, time_out) -> str | None:
    if status == 'Time Out':
        if _dt_empty(time_in) or _dt_empty(time_out):
            return "Time Out requires both time in and time out."
    elif status == 'Time In':
        if _dt_empty(time_in):
            return "Time In requires time in."
    elif status == 'Partial (No Time Out)':
        if _dt_empty(time_in):
            return "Partial (No Time Out) requires time in."
    return None


def update_session_attendance_record(record_id: int, session_id: str, fields: dict) -> tuple[bool, str, dict | None]:
    """
    Update one attendance row for a closed session. *fields* may include:
    status, time_in, time_out, fine, fine_reason, name, course, year, section.
    Only keys present in *fields* are applied (None clears optional text fields where safe).
    """
    ok, msg = assert_session_closed_for_attendance_edit(session_id)
    if not ok:
        return False, msg, None

    with get_db() as conn:
        cur = _cur(conn)
        cur.execute(
            """SELECT id, recorded_at, name, student_number, course, year, section, session_id,
                      status, fine, fine_reason, time_in, time_out
               FROM attendance_records WHERE id = %s AND session_id = %s""",
            (record_id, session_id),
        )
        row = cur.fetchone()
        if not row:
            return False, "Attendance record not found for this session.", None

        cur_status = row['status']
        cur_ti = row.get('time_in')
        cur_to = row.get('time_out')
        cur_fine = int(row.get('fine') or 0)
        cur_reason = row.get('fine_reason') or ''

        if 'status' in fields and fields['status'] is not None:
            st = str(fields['status']).strip()
            if st not in ALLOWED_EDIT_STATUSES:
                return False, f"Invalid status. Allowed: {', '.join(sorted(ALLOWED_EDIT_STATUSES))}.", None
            cur_status = st
        if 'time_in' in fields:
            cur_ti = normalize_attendance_datetime(fields['time_in']) if fields['time_in'] is not None else None
        if 'time_out' in fields:
            v = fields['time_out']
            cur_to = normalize_attendance_datetime(v) if v not in (None, '') else None
        if 'fine' in fields and fields['fine'] is not None:
            try:
                cur_fine = max(0, int(fields['fine']))
            except (TypeError, ValueError):
                return False, "Fine must be a non-negative integer.", None
        if 'fine_reason' in fields:
            cur_reason = (fields['fine_reason'] if fields['fine_reason'] is not None else '') or ''

        name = row.get('name')
        course = row.get('course')
        year = row.get('year')
        section = row.get('section')
        if 'name' in fields and fields['name'] is not None:
            name = str(fields['name']).strip() or name
        if 'course' in fields and fields['course'] is not None:
            course = str(fields['course']).strip() or course
        if 'year' in fields and fields['year'] is not None:
            year = str(fields['year']).strip() or year
        if 'section' in fields and fields['section'] is not None:
            section = str(fields['section']).strip() or section

        if cur_status == 'Absent':
            cur_ti, cur_to = None, None
        elif cur_status == 'Partial (No Time Out)':
            cur_to = None
        elif cur_status == 'Time In':
            cur_to = None

        err = _validate_attendance_edit_payload(cur_status, cur_ti, cur_to)
        if err:
            return False, err, None

        cur.execute(
            """UPDATE attendance_records SET
                   status = %s, time_in = %s, time_out = %s, fine = %s, fine_reason = %s,
                   name = %s, course = %s, year = %s, section = %s
               WHERE id = %s AND session_id = %s""",
            (cur_status, cur_ti, cur_to, cur_fine, cur_reason, name, course, year, section,
             record_id, session_id),
        )

        cur.execute(
            """SELECT id, recorded_at as datetime, name, student_number, course, year, section,
                      session_id, status, fine, fine_reason, time_in, time_out
               FROM attendance_records WHERE id = %s""",
            (record_id,),
        )
        out = dict(cur.fetchone())

    return True, "Attendance updated.", out


def add_manual_attendance_record(
    session_id: str,
    student_data: dict,
    status: str,
    time_in,
    time_out,
    fine: int = 0,
    fine_reason: str = '',
) -> tuple[bool, str, dict | None]:
    """
    Insert a new attendance row for a student who has no row yet for this closed session.
    """
    ok, msg = assert_session_closed_for_attendance_edit(session_id)
    if not ok:
        return False, msg, None

    st = str(status or '').strip()
    if st not in ALLOWED_EDIT_STATUSES:
        return False, f"Invalid status. Allowed: {', '.join(sorted(ALLOWED_EDIT_STATUSES))}.", None

    ti = normalize_attendance_datetime(time_in)
    to_val = normalize_attendance_datetime(time_out)
    if st == 'Absent':
        ti, to_val = None, None
    elif st == 'Partial (No Time Out)':
        to_val = None
    elif st == 'Time In':
        to_val = None

    err = _validate_attendance_edit_payload(st, ti, to_val)
    if err:
        return False, err, None

    try:
        fine_i = max(0, int(fine or 0))
    except (TypeError, ValueError):
        return False, "Fine must be a non-negative integer.", None

    sn = str(student_data.get('student_number', '')).strip()
    if not sn:
        return False, "Student number is required.", None

    now = ph_now().strftime('%Y-%m-%d %H:%M:%S')
    recorded_at = ti or now

    with get_db() as conn:
        cur = _cur(conn)
        cur.execute(
            "SELECT id FROM attendance_records WHERE session_id = %s AND student_number = %s LIMIT 1",
            (session_id, sn),
        )
        if cur.fetchone():
            return False, "This student already has an attendance row for this session.", None

        cur.execute(
            """INSERT INTO attendance_records
               (recorded_at, name, student_number, course, year, section, session_id,
                status, fine, fine_reason, time_in, time_out)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
               RETURNING id""",
            (
                recorded_at,
                student_data.get('name', ''),
                sn,
                student_data.get('course', ''),
                student_data.get('year', ''),
                student_data.get('section', ''),
                session_id,
                st,
                fine_i,
                fine_reason or '',
                ti,
                to_val,
            ),
        )
        new_id = cur.fetchone()['id']
        cur.execute(
            """SELECT id, recorded_at as datetime, name, student_number, course, year, section,
                      session_id, status, fine, fine_reason, time_in, time_out
               FROM attendance_records WHERE id = %s""",
            (new_id,),
        )
        out = dict(cur.fetchone())

    return True, "Attendance record added.", out


def get_session_stats(session_id: str) -> dict:
    """Get attendance statistics for a specific session using SQL aggregation."""
    with get_db() as conn:
        cur = _cur(conn)
        cur.execute("""
            SELECT COUNT(*) AS total,
                   COUNT(*) FILTER (
                       WHERE status = 'Time In'
                         AND (time_out IS NULL OR TRIM(COALESCE(time_out, '')) = '')
                   ) AS time_in,
                   COUNT(*) FILTER (
                       WHERE (time_out IS NOT NULL AND TRIM(COALESCE(time_out, '')) != '')
                          OR status = 'Time Out'
                   ) AS time_out,
                   COUNT(*) FILTER (WHERE status = 'Absent') AS absent,
                   COALESCE(SUM(fine), 0) AS total_fines
            FROM attendance_records WHERE session_id = %s
        """, (session_id,))
        agg = cur.fetchone()

        cur.execute("""
            SELECT CONCAT(COALESCE(course,''), ' ', COALESCE(year,''), '-', COALESCE(section,'')) AS key,
                   COUNT(*) AS cnt
            FROM attendance_records WHERE session_id = %s
            GROUP BY course, year, section
        """, (session_id,))
        by_course = {r['key']: r['cnt'] for r in cur.fetchall()}

    records = get_attendance_records(session_id=session_id)

    return {
        'total_present': agg['total'],
        'time_in_count': agg['time_in'],
        'time_out_count': agg['time_out'],
        'absent_count': agg['absent'],
        'total_fines': agg['total_fines'],
        'by_course': by_course,
        'records': records,
    }


def create_sample_master_list(filepath: str = None) -> str:
    """Create a sample Excel master list with student data for testing."""
    if filepath is None:
        filepath = os.path.join(EXCEL_DIR, 'sample_master_list.xlsx')

    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    headers = ['Name', 'Student Number', 'Course', 'Year', 'Section', 'School Year']
    widths = [30, 20, 15, 8, 10, 15]
    _style_header_row(ws, headers, widths)

    sample_students = [
        ['Juan Dela Cruz', '2024-00001', 'BSCS', '3', 'A', '2024-2025'],
        ['Maria Santos', '2024-00002', 'BSCS', '3', 'A', '2024-2025'],
        ['Jose Rizal Jr.', '2024-00003', 'BSIT', '2', 'B', '2024-2025'],
        ['Ana Reyes', '2024-00004', 'BSIT', '2', 'B', '2024-2025'],
        ['Carlos Garcia', '2024-00005', 'BSCS', '1', 'A', '2024-2025'],
        ['Elena Cruz', '2024-00006', 'BSCE', '4', 'A', '2024-2025'],
        ['Miguel Torres', '2024-00007', 'BSCS', '3', 'B', '2024-2025'],
        ['Sofia Bautista', '2024-00008', 'BSIT', '1', 'A', '2024-2025'],
        ['Rafael Mendoza', '2024-00009', 'BSCS', '2', 'A', '2024-2025'],
        ['Isabella Flores', '2024-00010', 'BSCE', '3', 'A', '2024-2025'],
    ]

    for row_idx, student in enumerate(sample_students, 2):
        for col_idx, value in enumerate(student, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')

    try:
        wb.save(filepath)
    finally:
        wb.close()
    return filepath


def clear_attendance_records() -> int:
    """Clear all attendance records from the database. Returns count deleted."""
    with get_db() as conn:
        cur = _cur(conn)
        cur.execute("SELECT COUNT(*) AS cnt FROM attendance_records")
        count = cur.fetchone()['cnt']
        cur.execute("DELETE FROM attendance_records")
    return count


def clear_session_records(session_id: str) -> int:
    """Clear attendance records for a specific session. Returns count deleted."""
    with get_db() as conn:
        cur = _cur(conn)
        cur.execute("SELECT COUNT(*) AS cnt FROM attendance_records WHERE session_id = %s",
                    (session_id,))
        count = cur.fetchone()['cnt']
        cur.execute("DELETE FROM attendance_records WHERE session_id = %s", (session_id,))
    return count
