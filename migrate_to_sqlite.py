#!/usr/bin/env python3
"""
One-time migration script to import existing JSON and Excel data into SQLite.
Run this once if you have data in sessions.json, student_registry.json, or attendance_log.xlsx.

Usage: python migrate_to_sqlite.py
"""
import json
import os
import sys
from datetime import datetime

# Add project root to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from config import (SESSIONS_FILE, STUDENT_REGISTRY_FILE, ATTENDANCE_LOG_FILE,
                    EXCEL_DIR, DATABASE_PATH)
from db import init_db, get_db


def migrate_students():
    """Import students from student_registry.json"""
    if not os.path.exists(STUDENT_REGISTRY_FILE):
        print("  No student_registry.json found, skipping.")
        return 0

    with open(STUDENT_REGISTRY_FILE, 'r') as f:
        registry = json.load(f)

    if not registry:
        print("  student_registry.json is empty, skipping.")
        return 0

    count = 0
    with get_db() as conn:
        for student_number, data in registry.items():
            try:
                conn.execute(
                    """INSERT OR REPLACE INTO students (student_number, name, course, year, section)
                       VALUES (?, ?, ?, ?, ?)""",
                    (
                        str(student_number),
                        str(data.get('name', '')),
                        str(data.get('course', '')),
                        str(data.get('year', '')),
                        str(data.get('section', '')),
                    ),
                )
                count += 1
            except Exception as e:
                print(f"  Warning: Could not import student {student_number}: {e}")
    return count


def migrate_sessions():
    """Import sessions from sessions.json"""
    if not os.path.exists(SESSIONS_FILE):
        print("  No sessions.json found, skipping.")
        return 0

    with open(SESSIONS_FILE, 'r') as f:
        sessions = json.load(f)

    if not sessions:
        print("  sessions.json is empty, skipping.")
        return 0

    count = 0
    with get_db() as conn:
        for session_id, data in sessions.items():
            try:
                required_year = data.get('required_year', [])
                if isinstance(required_year, list):
                    required_year = json.dumps(required_year)
                elif not isinstance(required_year, str):
                    required_year = '[]'

                conn.execute(
                    """INSERT OR REPLACE INTO sessions
                       (session_id, subject, teacher, notes, created_at, expires_at,
                        is_active, required_course, required_year, required_section)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (
                        session_id,
                        data.get('subject', ''),
                        data.get('teacher', ''),
                        data.get('notes', ''),
                        data.get('created_at', ''),
                        data.get('expires_at', ''),
                        1 if data.get('is_active', True) else 0,
                        data.get('required_course', ''),
                        required_year,
                        data.get('required_section', ''),
                    ),
                )

                scanned = data.get('scanned_students', {})
                if isinstance(scanned, list):
                    scanned = {s: {'status': 'in', 'time_in': data.get('created_at', ''), 'time_out': None, 'fine': 0, 'fine_reason': ''} for s in scanned}

                for student_number, info in scanned.items():
                    if isinstance(info, str):
                        continue
                    conn.execute(
                        """INSERT OR REPLACE INTO session_scans
                           (session_id, student_number, status, time_in, time_out, fine, fine_reason)
                           VALUES (?, ?, ?, ?, ?, ?, ?)""",
                        (
                            session_id,
                            student_number,
                            info.get('status', 'in'),
                            info.get('time_in', ''),
                            info.get('time_out'),
                            info.get('fine', 0),
                            info.get('fine_reason', '') or '',
                        ),
                    )
                count += 1
            except Exception as e:
                print(f"  Warning: Could not import session {session_id}: {e}")
    return count


def migrate_attendance():
    """Import attendance records from attendance_log.xlsx"""
    if not os.path.exists(ATTENDANCE_LOG_FILE):
        print("  No attendance_log.xlsx found, skipping.")
        return 0

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("  openpyxl not installed, skipping Excel migration.")
        return 0

    wb = load_workbook(ATTENDANCE_LOG_FILE, read_only=True)
    count = 0

    try:
        with get_db() as conn:
            for sheet_name in wb.sheetnames:
                if sheet_name == 'Summary':
                    continue
                ws = wb[sheet_name]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not row[0]:
                        continue
                    try:
                        conn.execute(
                            """INSERT INTO attendance_records
                               (recorded_at, name, student_number, course, year, section,
                                session_id, status, fine, fine_reason)
                               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                            (
                                str(row[0]) if row[0] else datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                                row[1] or '',
                                str(row[2]) if row[2] else '',
                                row[3] or '',
                                row[4] or '',
                                row[5] or '',
                                row[6] or '',
                                row[7] or 'Present',
                                row[8] if len(row) > 8 and row[8] is not None else 0,
                                row[9] if len(row) > 9 else '',
                            ),
                        )
                        count += 1
                    except Exception as e:
                        print(f"  Warning: Could not import row: {e}")
    finally:
        wb.close()

    return count


def main():
    print("QR Attendance System - SQLite Migration")
    print("=" * 50)

    os.makedirs(EXCEL_DIR, exist_ok=True)
    init_db()

    if os.path.exists(DATABASE_PATH):
        with get_db() as conn:
            cursor = conn.execute("SELECT COUNT(*) FROM students")
            sc = cursor.fetchone()[0]
            cursor = conn.execute("SELECT COUNT(*) FROM sessions")
            ss = cursor.fetchone()[0]
            cursor = conn.execute("SELECT COUNT(*) FROM attendance_records")
            ac = cursor.fetchone()[0]
        if sc > 0 or ss > 0 or ac > 0:
            print("\nDatabase already has data:")
            print(f"  Students: {sc}, Sessions: {ss}, Attendance records: {ac}")
            r = input("Overwrite/merge with imported data? [y/N]: ").strip().lower()
            if r != 'y':
                print("Aborted.")
                return

    print("\nMigrating students...")
    n_students = migrate_students()
    print(f"  Imported {n_students} students.")

    print("\nMigrating sessions...")
    n_sessions = migrate_sessions()
    print(f"  Imported {n_sessions} sessions.")

    print("\nMigrating attendance records from Excel...")
    n_records = migrate_attendance()
    print(f"  Imported {n_records} attendance records.")

    print("\n" + "=" * 50)
    print("Migration complete!")
    print(f"Database: {DATABASE_PATH}")
    print("\nYou can now safely archive or delete:")
    print(f"  - {SESSIONS_FILE}")
    print(f"  - {STUDENT_REGISTRY_FILE}")
    print(f"  - {ATTENDANCE_LOG_FILE} (optional, data is now in SQLite)")


if __name__ == "__main__":
    main()
